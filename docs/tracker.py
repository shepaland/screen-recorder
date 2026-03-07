#!/usr/bin/env python3
"""
tracker.py — CLI-скрипт для управления tracker.xlsx
Использование:
  python tracker.py tasks list
  python tracker.py tasks add --type Bug --title "Текст" --desc "Описание" --priority High
  python tracker.py tasks update T-001 --status "In Progress"
  python tracker.py tests list
  python tracker.py tests add --feature "Авторизация" --title "Текст" --steps "1.шаг" --expected "Результат" --task T-001
  python tracker.py tests update TC-001 --status Pass
"""

import argparse
import sys
import os
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TRACKER_PATH = os.path.join(os.path.dirname(__file__), "tracker.xlsx")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TASK_STATUS   = ["Open", "In Progress", "Done", "Cancelled"]
TASK_PRIORITY = ["Low", "Medium", "High", "Critical"]
TASK_TYPES    = ["Task", "Bug"]
TC_STATUS     = ["Pending", "Pass", "Fail"]

STATUS_COLORS_TC   = {"Pass": "E2EFDA", "Fail": "FFE0E0", "Pending": "FFF2CC"}
TYPE_COLORS_TASK   = {"Bug": "FFF0F0", "Task": "F0F6FF"}


def _wb():
    return load_workbook(TRACKER_PATH)


def _save(wb):
    wb.save(TRACKER_PATH)


def _style_row(ws, row_num, bg_hex):
    for cell in ws[row_num]:
        cell.border = BORDER
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=True,
                                   horizontal="center" if cell.column in (1,2,5,6,8,9) else "left")
        if bg_hex:
            cell.fill = PatternFill("solid", start_color=bg_hex)


def _next_id(ws, prefix):
    """Generate next ID like T-004 or TC-005"""
    max_num = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] and str(row[0]).startswith(prefix):
            try:
                num = int(str(row[0]).split("-")[1])
                max_num = max(max_num, num)
            except (IndexError, ValueError):
                pass
    return f"{prefix}-{str(max_num + 1).zfill(3)}"


def _print_table(headers, rows):
    if not rows:
        print("  (нет записей)")
        return
    widths = [len(h) for h in headers]
    for row in rows:
        for i, cell in enumerate(row):
            widths[i] = max(widths[i], len(str(cell or "")))
    fmt = "  " + "  ".join(f"{{:<{w}}}" for w in widths)
    print(fmt.format(*headers))
    print("  " + "  ".join("-" * w for w in widths))
    for row in rows:
        print(fmt.format(*[str(c or "") for c in row]))


# ── TASKS ─────────────────────────────────────────────────────────────────────

def tasks_list(args):
    wb = _wb()
    ws = wb["Tasks"]
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        if args.status and row[5] != args.status:
            continue
        if args.type and row[1] != args.type:
            continue
        rows.append([row[0], row[1], (row[2] or "")[:40], row[4], row[5], row[6] or "—"])
    _print_table(["ID", "Тип", "Заголовок", "Приоритет", "Статус", "Исполнитель"], rows)


def tasks_add(args):
    wb = _wb()
    ws = wb["Tasks"]
    today = date.today().strftime("%d.%m.%Y")
    new_id = _next_id(ws, "T")
    row_num = ws.max_row + 1
    row_data = [new_id, args.type, args.title, args.desc or "", args.priority, "Open", args.assignee or "—", today, today]
    for col, val in enumerate(row_data, 1):
        ws.cell(row=row_num, column=col, value=val)
    ws.row_dimensions[row_num].height = 20
    _style_row(ws, row_num, TYPE_COLORS_TASK.get(args.type, "FFFFFF"))
    _save(wb)
    print(f"✅ Добавлено: {new_id} — {args.title}")


def tasks_update(args):
    wb = _wb()
    ws = wb["Tasks"]
    col_map = {"type": 2, "title": 3, "desc": 4, "priority": 5, "status": 6, "assignee": 7}
    today = date.today().strftime("%d.%m.%Y")
    found = False
    for row in ws.iter_rows(min_row=3):
        if row[0].value == args.id:
            found = True
            for field, col_idx in col_map.items():
                val = getattr(args, field, None)
                if val:
                    row[col_idx - 1].value = val
            row[8].value = today  # updated
            # Update row color if type changed
            if args.type:
                bg = TYPE_COLORS_TASK.get(args.type, "FFFFFF")
                _style_row(ws, row[0].row, bg)
            break
    if not found:
        print(f"❌ Задача {args.id} не найдена")
        sys.exit(1)
    _save(wb)
    print(f"✅ Обновлено: {args.id}")


# ── TEST CASES ────────────────────────────────────────────────────────────────

def tests_list(args):
    wb = _wb()
    ws = wb["TestCases"]
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        if args.status and row[7] != args.status:
            continue
        if args.feature and args.feature.lower() not in str(row[1] or "").lower():
            continue
        rows.append([row[0], row[1], (row[2] or "")[:35], row[7], row[8] or "—", row[9] or "—"])
    _print_table(["ID", "Фича", "Заголовок", "Статус", "Задача", "Дата прогона"], rows)


def tests_add(args):
    wb = _wb()
    ws = wb["TestCases"]
    today = date.today().strftime("%d.%m.%Y")
    new_id = _next_id(ws, "TC")
    row_num = ws.max_row + 1
    row_data = [new_id, args.feature, args.title, args.preconditions or "",
                args.steps or "", args.expected or "", "", "Pending",
                args.task or "—", today]
    for col, val in enumerate(row_data, 1):
        ws.cell(row=row_num, column=col, value=val)
    ws.row_dimensions[row_num].height = 50
    _style_row(ws, row_num, STATUS_COLORS_TC["Pending"])
    _save(wb)
    print(f"✅ Добавлено: {new_id} — {args.title}")


def tests_update(args):
    wb = _wb()
    ws = wb["TestCases"]
    col_map = {"feature": 2, "title": 3, "preconditions": 4, "steps": 5,
               "expected": 6, "actual": 7, "status": 8, "task": 9}
    today = date.today().strftime("%d.%m.%Y")
    found = False
    for row in ws.iter_rows(min_row=3):
        if row[0].value == args.id:
            found = True
            for field, col_idx in col_map.items():
                val = getattr(args, field, None)
                if val:
                    row[col_idx - 1].value = val
            row[9].value = today  # date
            if args.status:
                bg = STATUS_COLORS_TC.get(args.status, "FFFFFF")
                _style_row(ws, row[0].row, bg)
            break
    if not found:
        print(f"❌ Тест-кейс {args.id} не найден")
        sys.exit(1)
    _save(wb)
    print(f"✅ Обновлено: {args.id}")


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(description="Трекер задач и тест-кейсов (tracker.xlsx)")
    sub = p.add_subparsers(dest="entity", required=True)

    # tasks
    pt = sub.add_parser("tasks", help="Управление задачами и багами")
    ts = pt.add_subparsers(dest="action", required=True)

    tl = ts.add_parser("list", help="Список задач")
    tl.add_argument("--status", choices=TASK_STATUS)
    tl.add_argument("--type", choices=TASK_TYPES)
    tl.set_defaults(func=tasks_list)

    ta = ts.add_parser("add", help="Добавить задачу")
    ta.add_argument("--type", choices=TASK_TYPES, default="Task")
    ta.add_argument("--title", required=True)
    ta.add_argument("--desc")
    ta.add_argument("--priority", choices=TASK_PRIORITY, default="Medium")
    ta.add_argument("--assignee")
    ta.set_defaults(func=tasks_add)

    tu = ts.add_parser("update", help="Обновить задачу")
    tu.add_argument("id", help="ID задачи, например T-001")
    tu.add_argument("--type", choices=TASK_TYPES)
    tu.add_argument("--title")
    tu.add_argument("--desc")
    tu.add_argument("--priority", choices=TASK_PRIORITY)
    tu.add_argument("--status", choices=TASK_STATUS)
    tu.add_argument("--assignee")
    tu.set_defaults(func=tasks_update)

    # tests
    pc = sub.add_parser("tests", help="Управление тест-кейсами")
    cs = pc.add_subparsers(dest="action", required=True)

    cl = cs.add_parser("list", help="Список тест-кейсов")
    cl.add_argument("--status", choices=TC_STATUS)
    cl.add_argument("--feature")
    cl.set_defaults(func=tests_list)

    ca = cs.add_parser("add", help="Добавить тест-кейс")
    ca.add_argument("--feature", required=True)
    ca.add_argument("--title", required=True)
    ca.add_argument("--preconditions")
    ca.add_argument("--steps")
    ca.add_argument("--expected")
    ca.add_argument("--task", help="Связанная задача, например T-001")
    ca.set_defaults(func=tests_add)

    cu = cs.add_parser("update", help="Обновить тест-кейс")
    cu.add_argument("id", help="ID тест-кейса, например TC-001")
    cu.add_argument("--feature")
    cu.add_argument("--title")
    cu.add_argument("--preconditions")
    cu.add_argument("--steps")
    cu.add_argument("--expected")
    cu.add_argument("--actual")
    cu.add_argument("--status", choices=TC_STATUS)
    cu.add_argument("--task")
    cu.set_defaults(func=tests_update)

    args = p.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
